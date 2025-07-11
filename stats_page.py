import customtkinter as ctk
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from PIL import Image


class Stats_page(ctk.CTkFrame):
    def __init__(self, parent, controller=None):
        super().__init__(parent)
        self.show_frame = controller

        #Workbook instance
        workbook = Workbook()
        #Load existing workbook
        workbook = load_workbook("stats.xlsx")
        #Active worksheet
        worksheet = workbook.active
        
#Copilot in studio code---------------------------
        def get_stats():
            statistics = worksheet["B2":"P10"]
            rows = []
            for row in statistics:
                row_values = [str(cell.value) if cell.value is not None else "" for cell in row]
                rows.append("\t".join(row_values))  # Use tab for column separation
            label.configure(text="\n".join(rows))
#------------------------------------------------------ 


        #Stats Page content
        #Configure main grid
        self.grid_columnconfigure(0, weight=1) #Left column
        self.grid_columnconfigure(1, weight=1) #Right column
        self.grid_rowconfigure(0, weight=1)  #Top navigation bar
        self.grid_rowconfigure(1, weight=1)  #Main page content

        #Navigation bar
        navigation_bar = ctk.CTkFrame(self, height=35, bg_color="#00A29E", fg_color="#00A29E")
        navigation_bar.grid(row=0, column=0, columnspan=2, sticky="ew")
        navigation_bar.grid_columnconfigure(0, weight=1)
        navigation_bar.grid_columnconfigure(1, weight=1)
        navigation_bar.grid_columnconfigure(2, weight=1)
        navigation_bar.grid_columnconfigure(3, weight=1)
        navigation_bar.grid_rowconfigure(0, weight=1)

        logo = ctk.CTkImage(light_image=Image.open("Images/Stats Logo.png"), size=(75, 75))
        logo_label = ctk.CTkLabel(navigation_bar, text="", image=logo)
        logo_label.grid(row=0, column=0, sticky="nsew")

        title_label = ctk.CTkLabel(navigation_bar, text="Stats2Drills", font=("ADLaM Display", 40), text_color="White")
        title_label.grid(row=0, column=1, columnspan=2, sticky="nsew")

        home_button = ctk.CTkButton(navigation_bar, width=120, height=40, text="Home", font=("ADLaM Display", 20), text_color="white", border_width=2, border_color="white", fg_color="#16CCCC", hover_color="#c7c7c7")
        home_button.grid(row=0, column=3, sticky="e")
        home_icon= ctk.CTkImage(light_image=Image.open("Images/white home icon.png"), dark_image=Image.open("Images/white home icon.png"), size=(35, 35))
        home_icon_button = ctk.CTkButton(navigation_bar, text="", image=home_icon, bg_color="#00A29E", fg_color="#00A29E", height=30, width=35, hover_color="#c7c7c7")
        home_icon_button.grid(row=0, column=4, sticky="w")
        #Configure home_button to display the home page
        home_button.configure(command=lambda: self.show_frame("Home_page"))
        home_icon_button.configure(command=lambda: self.show_frame("Home_page"))


        #Left Column, contains buttons and entry box
        left_column = ctk.CTkFrame(self, width=600, bg_color="#F2F2F2", fg_color="#F2F2F2")
        left_column.grid(row=1, column=0, sticky="nsew")
        left_column.grid_columnconfigure(0, weight=1)
        left_column.grid_rowconfigure(0, weight=1)
        left_column.grid_rowconfigure(1, weight=1)
        left_column.grid_rowconfigure(2, weight=1)
        left_column.grid_rowconfigure(3, weight=1)
        left_column.grid_rowconfigure(4, weight=1)

        analyse_stats_button = ctk.CTkButton(left_column, text="Analyse Statistics", font=("ADLaM Display", 25), text_color="white", fg_color="#FF7A53", hover_color="#c7c7c7", command=get_stats)
        analyse_stats_button.grid(row=0, column=0, sticky="n")

        instruction_label = ctk.CTkLabel(left_column, text="Enter Team Age", font=("Abadi", 15), text_color="Black")
        instruction_label.grid(row=1, column=0, sticky="s")
        team_age_entry = ctk.CTkEntry(left_column, corner_radius=0)
        team_age_entry.grid(row=2, column=0)
        age_description_label = ctk.CTkLabel(left_column, text="Must be either U8s, U10s, U12s, U14s, U16s, U18s or U20s", font=("Abadi", 13), text_color="Black", wraplength=150, justify="center")
        age_description_label.grid(row=3, column=0, sticky="n")

        generate_plan_button = ctk.CTkButton(left_column, text="Generate Training Plan", font=("ADLaM Display", 25), text_color="white", fg_color="#FF7A53", hover_color="#c7c7c7")
        generate_plan_button.grid(row=4, column=0)        


        #Right Column, display imported stats and analysed stats
        right_column = ctk.CTkFrame(self, width=600, bg_color="#F2F2F2", fg_color="#F2F2F2")
        right_column.grid(row=1, column=1, sticky="nsew")
        right_column.grid_columnconfigure(0, weight=1)
        right_column.grid_rowconfigure(0, weight=1)
        right_column.grid_rowconfigure(1, weight=1)

        label = ctk.CTkLabel(right_column)
        label.grid(row=0, column=0)

